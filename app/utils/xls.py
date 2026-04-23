import re
from typing import Optional, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def read_spreadsheet(path: str) -> pd.DataFrame:
    """Читает .xls/.xlsx/.ods в DataFrame с корректным engine."""
    lower_path = path.lower()
    if lower_path.endswith(".ods"):
        try:
            return pd.read_excel(path, header=None, engine="odf")
        except ImportError as exc:
            raise ValueError("Для чтения .ods нужно установить пакет odfpy") from exc
    return pd.read_excel(path, header=None)


def build_query_name(raw: str) -> str:

    base_name = raw.split("(", 1)[0].strip()
    if not base_name:
        return ""

    normalized = base_name.lower().replace("ё", "е")
    if "линдинет" not in normalized:
        return base_name
    
    variant_match = re.search(r"\b(20|30)\b", raw)
    if not variant_match:
        return base_name
    
    return f"{base_name} {variant_match.group(1)}"


def _find_header_columns(frame: pd.DataFrame) -> tuple[int, int, Optional[int]]:
    """Ищет строку заголовка и индексы столбцов названия и ШК."""
    header_row = header_col = None
    barcode_col: Optional[int] = None

    for r in range(frame.shape[0]):
        row_barcode_col: Optional[int] = None
        for c in range(frame.shape[1]):
            cell = str(frame.iat[r, c]).strip().lower()
            if header_col is None and "наименование товара" in cell:
                header_row, header_col = r, c
            if row_barcode_col is None and cell in {"шк", "штрих код", "штрихкод", "barcode"}:
                row_barcode_col = c
        if header_row is not None:
            barcode_col = row_barcode_col
            break

    if header_row is None or header_col is None:
        raise ValueError("Не найден столбец 'Наименование товара'")

    return header_row, header_col, barcode_col


def _find_product_code_column(frame: pd.DataFrame) -> tuple[int, int]:
    """Ищет строку заголовка и индекс столбца 'Код товара'."""
    for r in range(frame.shape[0]):
        for c in range(frame.shape[1]):
            cell = str(frame.iat[r, c]).strip().lower()
            if cell == "код товара":
                return r, c
    raise ValueError("Не найден столбец 'Код товара'")


def _normalize_barcode(value: object) -> str:
    """Нормализует штрих-код из Excel в строку без служебных символов."""
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    text = text.replace(" ", "")
    text = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def _normalize_product_code(value: object) -> str:
    """Нормализует код товара из Excel в строку цифр/символов без лишних пробелов."""
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    text = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def extract_product_codes_from_excel(path: str) -> list[dict]:
    """Извлекает из Excel коды товаров и связанные данные строки."""
    df = read_spreadsheet(path)
    header_row, code_col = _find_product_code_column(df)

    name_col: Optional[int] = None
    for c in range(df.shape[1]):
        cell = str(df.iat[header_row, c]).strip().lower()
        if "наименование товара" in cell:
            name_col = c
            break

    items: list[dict] = []
    seen: set[tuple[str, int]] = set()
    for row_idx in range(header_row + 1, df.shape[0]):
        product_code = _normalize_product_code(df.iat[row_idx, code_col])
        if not product_code:
            continue

        key = (product_code, row_idx)
        if key in seen:
            continue
        seen.add(key)

        name = ""
        if name_col is not None:
            raw_name = df.iat[row_idx, name_col]
            if raw_name is not None and not pd.isna(raw_name):
                name = str(raw_name).strip()

        items.append({
            "row_index": row_idx + 1,
            "product_code": product_code,
            "name": name,
        })

    return items


def _apteka_title(city_name: str) -> str:
    """Возвращает заголовок блока Apteka с учетом выбранного города."""
    normalized_city = (city_name or "").strip()
    return f"Apteka Ru - {normalized_city}" if normalized_city else "Apteka Ru"


def build_enriched_xlsx(
    path: str,
    out_path: str,
    items: list[dict],
    city_name: str = "",
    pharmacy_codes: Optional[list[str]] = None,
) -> None:
    """Дополняет исходную таблицу результатами парсинга и сохраняет как XLSX."""
    df = read_spreadsheet(path)

    header_row, header_col, barcode_col = _find_header_columns(df)

    def _is_empty(v) -> bool:
        if pd.isna(v):
            return True
        return str(v).strip() == ""

    def _find_list_start_row(frame: pd.DataFrame) -> Optional[int]:
        for row_idx in range(header_row + 1, frame.shape[0]):
            name_cell = frame.iat[row_idx, header_col]
            if _is_empty(name_cell):
                continue

            if header_col > 0:
                order_cell = frame.iat[row_idx, header_col - 1]
                if _is_empty(order_cell):
                    continue

                order_text = str(order_cell).strip()
                if re.fullmatch(r"\d+(?:\.0+)?", order_text):
                    return row_idx
            else:
                return row_idx

        return None

    def _key(name: str) -> str:
        return (name or "").strip().lower().replace("ё", "е")

    def _normalize_dosage(value: object) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip().lower().replace("ё", "е")
        if not text:
            return None

        def _format_number(number: float) -> str:
            return (f"{number:.6f}").rstrip("0").rstrip(".")

        def _normalize_part(number: float, unit: str) -> tuple[float, str]:
            unit = unit.lower()
            if unit == "мкг":
                return number / 1000, "мг"
            if unit in {"г", "гр"}:
                return number * 1000, "мг"
            return number, unit

        def _parentheses_depth(raw_text: str, idx: int) -> int:
            depth = 0
            for pos, ch in enumerate(raw_text):
                if pos >= idx:
                    break
                if ch == "(":
                    depth += 1
                elif ch == ")" and depth > 0:
                    depth -= 1
            return depth

        matches = list(re.finditer(r"\b(\d+(?:[\.,]\d+)?)\s*(мкг|мг|г|гр|мл|ме|iu|%)(?!\w)", text, flags=re.IGNORECASE))
        if not matches:
            return None

        min_depth = min(_parentheses_depth(text, m.start()) for m in matches)
        selected_matches = [m for m in matches if _parentheses_depth(text, m.start()) == min_depth]

        parsed_parts: list[tuple[float, str]] = []
        for m in selected_matches:
            raw_number = float(m.group(1))
            number, unit = _normalize_part(raw_number, m.group(2))
            parsed_parts.append((number, unit))

        if not parsed_parts:
            return None
        parsed_parts.sort(key=lambda part: (part[1], part[0]))
        parts = [f"{_format_number(number)} {unit}" for number, unit in parsed_parts]
        return " + ".join(parts)

    def _to_number(value: object) -> Optional[float]:
        if value is None:
            return None
        text_value = str(value).strip()
        if not text_value:
            return None

        normalized = text_value.replace(" ", "").replace(" ", "").replace(",", ".")
        normalized = re.sub(r"[^0-9.\-]", "", normalized)
        if not normalized:
            return None
        try:
            return float(normalized)
        except ValueError:
            return None

    def _extract_manufacturer_score(message: object) -> Optional[float]:
        if message is None:
            return None
        match = re.search(r"score производителя:\s*(\d+(?:[.,]\d+)?)%", str(message), flags=re.IGNORECASE)
        if not match:
            return None
        try:
            return float(match.group(1).replace(",", "."))
        except ValueError:
            return None

    def _extract_name_score(message: object) -> Optional[float]:
        if message is None:
            return None
        match = re.search(r"score названия:\s*(\d+(?:[.,]\d+)?)%", str(message), flags=re.IGNORECASE)
        if not match:
            return None
        try:
            return float(match.group(1).replace(",", "."))
        except ValueError:
            return None
    
    list_start_row = _find_list_start_row(df)

    if list_start_row is not None and list_start_row > header_row + 1:
        subheader_rows = range(header_row, list_start_row)
        for col_idx in range(df.shape[1]):
            parts: list[str] = []
            for row_idx in subheader_rows:
                cell_value = df.iat[row_idx, col_idx]
                if _is_empty(cell_value):
                    continue
                normalized = " ".join(str(cell_value).split())
                if normalized:
                    parts.append(normalized)

            if parts:
                df.iat[header_row, col_idx] = " ".join(parts)

        rows_to_remove = list(range(header_row + 1, list_start_row))
        if rows_to_remove:
            df = df.drop(index=rows_to_remove).reset_index(drop=True)

        list_start_row = header_row + 1

    product_code_col: Optional[int] = None
    try:
        _, product_code_col = _find_product_code_column(df)
    except ValueError:
        product_code_col = None

    main_extra_headers = [
        "Цена",
        "Отклонение от базовой цены",
        "Отклонение от закупочной цены",
        "Отклонение от нашего сайта",
    ]
    apteka_extra_headers = [
        "Оригинальные названия",
        "Найденный товар",
        "Сообщение",
    ]

    def _normalize_pharmacy_code(value: object) -> str:
        return str(value or "").strip().lower()

    def _pharmacy_title(code: str) -> str:
        if code == "apteka_ru":
            return _apteka_title(city_name)
        if code == "farmacia24":
            return "Губернские аптеки"
        return code.replace("_", " ").title()

    selected_codes = [
        _normalize_pharmacy_code(code)
        for code in (pharmacy_codes or [])
        if _normalize_pharmacy_code(code)
    ]
    if selected_codes:
        pharmacy_codes = []
        for code in selected_codes:
            if code not in pharmacy_codes:
                pharmacy_codes.append(code)
    else:
        pharmacy_codes = []
        for item in items:
            code = _normalize_pharmacy_code(item.get("source_pharmacy")) or "apteka_ru"
            if code not in pharmacy_codes:
                pharmacy_codes.append(code)
        if not pharmacy_codes:
            pharmacy_codes = ["apteka_ru"]

    found_name_col_by_code: dict[str, int] = {}
    found_names_start_col = header_col + 1
    for idx, code in enumerate(pharmacy_codes):
        col_idx = found_names_start_col + idx
        df.insert(col_idx, f"__found_name__{code}", None)
        found_name_col_by_code[code] = col_idx
        df.iat[header_row, col_idx] = _pharmacy_title(code)

    inserted_found_name_cols = len(pharmacy_codes)
    if inserted_found_name_cols:
        if barcode_col is not None and barcode_col > header_col:
            barcode_col += inserted_found_name_cols
        if product_code_col is not None and product_code_col > header_col:
            product_code_col += inserted_found_name_cols

    pharmacy_items: dict[str, list[dict]] = {code: [] for code in pharmacy_codes}
    for item in items:
        code = _normalize_pharmacy_code(item.get("source_pharmacy")) or "apteka_ru"
        pharmacy_items.setdefault(code, []).append(item)

    indexes_by_pharmacy: dict[str, dict[str, dict[str, list[dict]]]] = {}
    for code in pharmacy_codes:
        code_items = pharmacy_items.get(code, [])
        code_by_product: dict[str, list[dict]] = {}
        code_by_name: dict[str, list[dict]] = {}
        code_by_barcode: dict[str, list[dict]] = {}
        for item in code_items:
            product_code_key = str(item.get("input_product_code") or "").strip()
            if product_code_key:
                code_by_product.setdefault(product_code_key, []).append(item)

            key = _key(str(item.get("input_name") or ""))
            if key:
                code_by_name.setdefault(key, []).append(item)

            barcode_key = _normalize_barcode(item.get("input_barcode"))
            if barcode_key:
                code_by_barcode.setdefault(barcode_key, []).append(item)
        indexes_by_pharmacy[code] = {
            "product": code_by_product,
            "name": code_by_name,
            "barcode": code_by_barcode,
        }

    def _column_is_empty(col_idx: int) -> bool:
        if col_idx >= df.shape[1]:
            return True
        for row_idx in range(header_row, df.shape[0]):
            if not _is_empty(df.iat[row_idx, col_idx]):
                return False
        return True

    insert_col = header_col + 1
    parsed_block_width = len(main_extra_headers) * len(pharmacy_codes)
    while True:
        block_is_free = True
        for offset in range(parsed_block_width):
            if not _column_is_empty(insert_col + offset):
                block_is_free = False
                break
        if block_is_free:
            break
        insert_col += 1

    required_cols = insert_col + parsed_block_width
    while df.shape[1] < required_cols:
        df[df.shape[1]] = None

    block_start_by_code: dict[str, int] = {}
    for idx, code in enumerate(pharmacy_codes):
        start_col = insert_col + idx * len(main_extra_headers)
        block_start_by_code[code] = start_col
        for offset, name in enumerate(main_extra_headers):
            df.iat[header_row, start_col + offset] = name

    apteka_rows_by_code: dict[str, dict[int, list[object]]] = {code: {} for code in pharmacy_codes}
    warning_rows_by_code: dict[str, set[int]] = {code: set() for code in pharmacy_codes}
    no_info_rows_by_code: dict[str, set[int]] = {code: set() for code in pharmacy_codes}

    base_price_col: Optional[int] = None
    purchase_price_col: Optional[int] = None
    site_price_col: Optional[int] = None
    fallback_site_price_col: Optional[int] = None
    for col_idx in range(df.shape[1]):
        header_value = " ".join(str(df.iat[header_row, col_idx]).strip().lower().split())
        if base_price_col is None and "цена базовая" in header_value:
            base_price_col = col_idx
        if purchase_price_col is None and "цена закуп" in header_value:
            purchase_price_col = col_idx
        if site_price_col is None and "фг- it-к" in header_value:
            site_price_col = col_idx
        if fallback_site_price_col is None and "цена фг" in header_value:
            fallback_site_price_col = col_idx
        if (
            base_price_col is not None
            and purchase_price_col is not None
            and site_price_col is not None
        ):
            break

    if site_price_col is None:
        site_price_col = fallback_site_price_col

    base_markup_formula_rows_by_code: dict[str, list[int]] = {code: [] for code in pharmacy_codes}
    purchase_markup_formula_rows_by_code: dict[str, list[int]] = {code: [] for code in pharmacy_codes}
    site_markup_formula_rows_by_code: dict[str, list[int]] = {code: [] for code in pharmacy_codes}

    for r in range(header_row + 1, df.shape[0]):
        raw = df.iat[r, header_col]
        if _is_empty(raw):
            continue

        raw_text = str(raw)
        query_name = build_query_name(raw_text)
        if not query_name:
            continue

        query_qty, query_qty_is_sum = extract_qty_from_xls_row(raw_text)
        query_qty_pack = extract_qty_pack_format(raw_text)
        query_dosage = _normalize_dosage(extract_dosage_from_xls_row(raw_text))
        query_barcode = _normalize_barcode(df.iat[r, barcode_col]) if barcode_col is not None else ""
        query_product_code = str(df.iat[r, product_code_col]).strip() if product_code_col is not None and not _is_empty(df.iat[r, product_code_col]) else ""

        def _candidate_dosage(candidate: dict) -> Optional[str]:
            return _normalize_dosage(candidate.get("input_dosage"))

        for code in pharmacy_codes:
            code_indexes = indexes_by_pharmacy.get(code, {})
            code_by_product = code_indexes.get("product", {})
            code_by_barcode = code_indexes.get("barcode", {})
            code_by_name = code_indexes.get("name", {})

            candidates = code_by_product.get(query_product_code, []) if query_product_code else []
            if not candidates and query_barcode:
                candidates = code_by_barcode.get(query_barcode, [])
            if not candidates:
                candidates = code_by_name.get(_key(query_name), [])
            if not candidates:
                no_info_rows_by_code[code].add(r)
                continue

            qty_matched: list[dict] = []
            if query_qty_pack is not None:
                qty_matched = [
                    c
                    for c in candidates
                    if extract_qty_pack_format(str(c.get("input_qty") or "")) == query_qty_pack
                ]
                if not qty_matched and query_qty is not None:
                    qty_matched = [c for c in candidates if c.get("input_qty") == query_qty]
                if not qty_matched:
                    no_info_rows_by_code[code].add(r)
                    continue
            elif query_qty is not None:
                qty_matched = [c for c in candidates if c.get("input_qty") == query_qty]
                if not qty_matched:
                    no_info_rows_by_code[code].add(r)
                    continue
            else:
                qty_matched = [c for c in candidates if c.get("input_qty") is None]
                if not qty_matched:
                    qty_matched = candidates

            if query_dosage is not None:
                dosage_matched = [c for c in qty_matched if _candidate_dosage(c) == query_dosage]
                if not dosage_matched:
                    no_info_rows_by_code[code].add(r)
                    continue
                item = dosage_matched[0]
            else:
                no_dosage = [c for c in qty_matched if _candidate_dosage(c) is None]
                item = no_dosage[0] if no_dosage else qty_matched[0]

            parsed_price = item.get("price", "")
            if _is_empty(parsed_price):
                no_info_rows_by_code[code].add(r)

            block_start = block_start_by_code[code]
            found_name_col = found_name_col_by_code.get(code)
            if found_name_col is not None:
                df.iat[r, found_name_col] = item.get("title", "")
            df.iat[r, block_start] = parsed_price
            df.iat[r, block_start + 1] = ""
            df.iat[r, block_start + 2] = ""
            df.iat[r, block_start + 3] = ""

            parsed_price_num = _to_number(parsed_price)
            if base_price_col is not None:
                base_price = _to_number(df.iat[r, base_price_col])
                if base_price is not None and parsed_price_num is not None:
                    base_markup_formula_rows_by_code[code].append(r)
            if purchase_price_col is not None:
                purchase_price = _to_number(df.iat[r, purchase_price_col])
                if purchase_price is not None and parsed_price_num is not None:
                    purchase_markup_formula_rows_by_code[code].append(r)
            if site_price_col is not None:
                site_price = _to_number(df.iat[r, site_price_col])
                if site_price is not None and parsed_price_num is not None:
                    site_markup_formula_rows_by_code[code].append(r)

            message_text = str(item.get("message", ""))
            message_lower = message_text.lower()
            dosage_no_data = (
                "совпадение дозировки: нет данных" in message_lower
                or "score дозировки: — (нет данных" in message_lower
            )
            qty_sum_warning = query_qty_is_sum or ("уточните цену на сайте, возможны неточности" in message_lower)
            manufacturer_score = _extract_manufacturer_score(message_text)
            manufacturer_warning = manufacturer_score is not None and 50 <= manufacturer_score < 80
            name_score_raw = item.get("name_score")
            try:
                name_score = float(name_score_raw) if name_score_raw is not None else None
            except (TypeError, ValueError):
                name_score = None
            if name_score is None:
                name_score = _extract_name_score(message_text)
            partial_name_warning = bool(item.get("partial_name_match")) or (name_score is not None and 50 < name_score < 80)
            full_name_match = name_score is not None and name_score >= 90
            full_manufacturer_match = manufacturer_score is None or manufacturer_score >= 80
            input_qty = item.get("input_qty")
            found_qty = item.get("found_qty")
            qty_exact_match = input_qty is None or input_qty == found_qty

            expected_dosage = _normalize_dosage(item.get("input_dosage"))
            found_dosage = _normalize_dosage(item.get("found_dosage"))
            dosage_similarity_percent = item.get("dosage_similarity_percent")
            dosage_warning = False
            if dosage_similarity_percent is not None:
                try:
                    similarity_value = int(dosage_similarity_percent)
                    dosage_warning = 50 <= similarity_value < 100
                except (TypeError, ValueError):
                    dosage_warning = False
            else:
                dosage_exact = dosage_no_data or expected_dosage is None or expected_dosage == found_dosage
                dosage_warning = not dosage_exact
            if dosage_no_data and full_name_match and full_manufacturer_match and qty_exact_match:
                dosage_warning = False
            
            warning_reasons: list[str] = []
            if qty_sum_warning:
                warning_reasons.append("количество требует проверки")
            if dosage_warning:
                warning_reasons.append("дозировка частично совпала")
            if manufacturer_warning:
                warning_reasons.append("производитель частично совпал")
            if partial_name_warning:
                warning_reasons.append("название частично совпало")

            if qty_sum_warning or dosage_warning or manufacturer_warning or partial_name_warning:
                warning_rows_by_code[code].add(r)

            short_warning_message = " / ".join(warning_reasons)
            message_for_extra_sheet = short_warning_message if short_warning_message and r not in no_info_rows_by_code[code] else ""
            apteka_rows_by_code[code][r] = [
                raw_text,
                item.get("title", ""),
                message_for_extra_sheet,
            ]

    wb = Workbook()
    ws = wb.active

    source_side = Side(style="thin", color="000000")
    parsed_side = Side(style="thin", color="000000")

    ROW_OFFSET = 1

    header_alignment = Alignment(vertical="top", wrap_text=True)
    content_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    price_numeric_columns = {
        *(block_start_by_code[code] for code in pharmacy_codes),
        *(idx for idx in [base_price_col, purchase_price_col, site_price_col] if idx is not None),
    }

    for row_idx in range(df.shape[0]):
        for col_idx in range(df.shape[1]):
            value = df.iat[row_idx, col_idx]
            cell = ws.cell(row=row_idx + 1 + ROW_OFFSET, column=col_idx + 1)
            if pd.isna(value):
                cell.value = ""
            elif row_idx > header_row and col_idx in price_numeric_columns:
                numeric_value = _to_number(value)
                cell.value = numeric_value if numeric_value is not None else value
            else:
                cell.value = value
            if row_idx == header_row:
                cell.alignment = header_alignment
            else:
                cell.alignment = content_alignment

    warning_fill = PatternFill(fill_type="solid", fgColor="FFE599")
    empty_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

    for code in pharmacy_codes:
        block_start = block_start_by_code[code]
        parsed_price_letter = get_column_letter(block_start + 1)

        if base_price_col is not None:
            base_price_letter = get_column_letter(base_price_col + 1)
            base_markup_col = block_start + 2
            for row_idx in base_markup_formula_rows_by_code[code]:
                excel_row = row_idx + 1 + ROW_OFFSET
                base_markup_cell = ws.cell(row=excel_row, column=base_markup_col)
                base_markup_cell.value = (
                    f"=IF(OR({parsed_price_letter}{excel_row}=0,{base_price_letter}{excel_row}=0),"
                    f"0,{parsed_price_letter}{excel_row}/{base_price_letter}{excel_row}-1)"
                )
                base_markup_cell.number_format = '0.00%'

        if purchase_price_col is not None:
            purchase_price_letter = get_column_letter(purchase_price_col + 1)
            purchase_markup_col = block_start + 3
            for row_idx in purchase_markup_formula_rows_by_code[code]:
                excel_row = row_idx + 1 + ROW_OFFSET
                purchase_markup_cell = ws.cell(row=excel_row, column=purchase_markup_col)
                purchase_markup_cell.value = (
                    f"=IF(OR({parsed_price_letter}{excel_row}=0,{purchase_price_letter}{excel_row}=0),"
                    f"0,{parsed_price_letter}{excel_row}/{purchase_price_letter}{excel_row}-1)"
                )
                purchase_markup_cell.number_format = '0.00%'

        if site_price_col is not None:
            site_price_letter = get_column_letter(site_price_col + 1)
            site_markup_col = block_start + 4
            for row_idx in site_markup_formula_rows_by_code[code]:
                excel_row = row_idx + 1 + ROW_OFFSET
                site_markup_cell = ws.cell(row=excel_row, column=site_markup_col)
                site_markup_cell.value = (
                    f"=IF(OR({parsed_price_letter}{excel_row}=0,{site_price_letter}{excel_row}=0),"
                    f"0,{parsed_price_letter}{excel_row}/{site_price_letter}{excel_row}-1)"
                )
                site_markup_cell.number_format = '0.00%'

        found_name_col = found_name_col_by_code.get(code)

        for row_idx in warning_rows_by_code[code]:
            if row_idx in no_info_rows_by_code[code]:
                continue
            excel_row = row_idx + 1 + ROW_OFFSET
            for col_idx in range(block_start, block_start + len(main_extra_headers)):
                ws.cell(row=excel_row, column=col_idx + 1).fill = warning_fill
            if found_name_col is not None:
                ws.cell(row=excel_row, column=found_name_col + 1).fill = warning_fill

        for row_idx in no_info_rows_by_code[code]:
            excel_row = row_idx + 1 + ROW_OFFSET
            for col_idx in range(block_start, block_start + len(main_extra_headers)):
                ws.cell(row=excel_row, column=col_idx + 1).fill = empty_fill
            if found_name_col is not None:
                ws.cell(row=excel_row, column=found_name_col + 1).fill = empty_fill

    source_min_col = 1
    source_max_col = insert_col
    parsed_min_col = insert_col + 1
    parsed_max_col = insert_col + parsed_block_width

    ws.merge_cells(
        start_row=1,
        start_column=source_min_col,
        end_row=1,
        end_column=source_max_col,
    )
    source_header_cell = ws.cell(row=1, column=source_min_col)
    source_header_cell.value = "Входящая информация"
    source_header_cell.alignment = Alignment(horizontal="center", vertical="center")
    source_header_cell.font = Font(size=18, bold=True)
    source_header_cell.fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    for code in pharmacy_codes:
        block_start = block_start_by_code[code] + 1
        block_end = block_start + len(main_extra_headers) - 1
        ws.merge_cells(
            start_row=1,
            start_column=block_start,
            end_row=1,
            end_column=block_end,
        )
        parsed_header_cell = ws.cell(row=1, column=block_start)
        parsed_header_cell.value = _pharmacy_title(code)
        parsed_header_cell.alignment = Alignment(horizontal="center", vertical="center")
        parsed_header_cell.font = Font(size=18, bold=True)
        parsed_header_cell.fill = PatternFill(fill_type="solid", fgColor="D0E0E3")

    def _max_line_len(value: object) -> int:
        text = "" if pd.isna(value) else str(value)
        if not text:
            return 0
        return max((len(line) for line in text.splitlines()), default=0)

    for col_idx in range(df.shape[1]):
        max_len = 0
        for row_idx in range(df.shape[0]):
            max_len = max(max_len, _max_line_len(df.iat[row_idx, col_idx]))

        width = min(max(max_len + 2, 6), 80)
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    def _apply_table_borders(min_row: int, max_row: int, min_col: int, max_col: int, side: Side) -> None:
        if min_row > max_row or min_col > max_col:
            return

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                is_top = r == min_row
                is_bottom = r == max_row
                is_left = c == min_col
                is_right = c == max_col
                has_inner_vertical = c > min_col

                cell = ws.cell(row=r + 1 + ROW_OFFSET, column=c + 1)
                border = cell.border
                cell.border = Border(
                    left=side if (is_left or has_inner_vertical) else border.left,
                    right=side if (is_right or has_inner_vertical) else border.right,
                    top=side if is_top else border.top,
                    bottom=side if is_bottom else border.bottom,
                )

    def _apply_outer_border(min_row: int, max_row: int, min_col: int, max_col: int, side: Side) -> None:
        if min_row > max_row or min_col > max_col:
            return

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                is_top = r == min_row
                is_bottom = r == max_row
                is_left = c == min_col
                is_right = c == max_col

                if not (is_top or is_bottom or is_left or is_right):
                    continue

                cell = ws.cell(row=r + 1 + ROW_OFFSET, column=c + 1)
                border = cell.border
                cell.border = Border(
                    left=side if is_left else border.left,
                    right=side if is_right else border.right,
                    top=side if is_top else border.top,
                    bottom=side if is_bottom else border.bottom,
                )

    last_row = df.shape[0] - 1
    if last_row >= header_row:
        subheader_min_row = header_row
        subheader_max_row = (list_start_row - 1) if list_start_row is not None else header_row
        if subheader_min_row <= subheader_max_row:
            _apply_outer_border(
                min_row=subheader_min_row,
                max_row=subheader_max_row,
                min_col=0,
                max_col=df.shape[1] - 1,
                side=source_side,
            )

        _apply_table_borders(
            min_row=header_row,
            max_row=last_row,
            min_col=0,
            max_col=insert_col - 1,
            side=source_side,
        )
        for code in pharmacy_codes:
            block_start = block_start_by_code[code]
            _apply_table_borders(
                min_row=header_row,
                max_row=last_row,
                min_col=block_start,
                max_col=block_start + len(main_extra_headers) - 1,
                side=parsed_side,
            )

    ws.title = "Итог"

    for code in pharmacy_codes:
        sheet_title = _pharmacy_title(code)[:31] or "Sheet"
        apteka_sheet = wb.create_sheet(sheet_title)
        apteka_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(apteka_extra_headers))

        apteka_header_cell = apteka_sheet.cell(row=1, column=1)
        apteka_header_cell.value = _pharmacy_title(code)
        apteka_header_cell.alignment = Alignment(horizontal="center", vertical="center")
        apteka_header_cell.font = Font(size=22, bold=True)
        apteka_header_cell.fill = PatternFill(fill_type="solid", fgColor="D0E0E3")

        for col_idx, header in enumerate(apteka_extra_headers, start=1):
            cell = apteka_sheet.cell(row=2, column=col_idx)
            cell.value = header
            cell.alignment = header_alignment
            cell.border = Border(left=parsed_side, right=parsed_side, top=parsed_side, bottom=parsed_side)

        apteka_row = 3
        pharmacy_rows = apteka_rows_by_code.get(code, {})
        for r in range(header_row + 1, df.shape[0]):
            raw_value = df.iat[r, header_col]
            original_name = "" if _is_empty(raw_value) else str(raw_value)
            values = pharmacy_rows.get(r, [original_name, "", ""])
            for col_idx, value in enumerate(values, start=1):
                cell = apteka_sheet.cell(row=apteka_row, column=col_idx)
                cell.value = value
                cell.alignment = content_alignment
                cell.border = Border(left=parsed_side, right=parsed_side, top=parsed_side, bottom=parsed_side)

            if r in no_info_rows_by_code.get(code, set()):
                for col_idx in range(1, len(apteka_extra_headers) + 1):
                    apteka_sheet.cell(row=apteka_row, column=col_idx).fill = empty_fill
            elif r in warning_rows_by_code.get(code, set()):
                for col_idx in range(1, len(apteka_extra_headers) + 1):
                    apteka_sheet.cell(row=apteka_row, column=col_idx).fill = warning_fill
            apteka_row += 1

        for offset in range(len(apteka_extra_headers)):
            target_letter = get_column_letter(offset + 1)
            if offset == 0:
                width = 40
            elif offset == 1:
                width = 40
            else:
                width = 50
            apteka_sheet.column_dimensions[target_letter].width = width

    wb.save(out_path)


def build_flat_xlsx(out_path: str, items: list[dict], city_name: str = "") -> None:
    """Сохраняет плоский список результатов в XLSX без исходной таблицы."""
    columns = ["input_name", "title", "price", "message"]
    df = pd.DataFrame(items, columns=columns)
    sheet_name = _apteka_title(city_name)[:31] or "Sheet1"
    df.to_excel(out_path, index=False, sheet_name=sheet_name)


def extract_qty_pack_format(text: str) -> Optional[str]:
    """Возвращает количество в формате `X+Y` (например, `21+7`) если оно есть в тексте."""
    if not text:
        return None

    match = re.search(
        r"(?:\bN\s*|№\s*)?(\d+)\s*(?:шт\.?)?\s*\+\s*(\d+)\b",
        str(text),
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    return f"{int(match.group(1))}+{int(match.group(2))}"


def extract_qty_from_xls_row(text: str) -> Tuple[Optional[int], bool]:
    """Возвращает количество из передаваемого текста"""
    if not text:
        return None, False

    normalized_text = str(text).lower().replace("ё", "е")

    # Упаковки вида "№28х3"/"N28x3" трактуем как 28 * 3 = 84.
    # Это используется в farmacia24, где количество иногда задается
    # через знак умножения вместо суммы.
    multiply_match = re.search(
        r"(?:\bN\s*|№\s*)(\d+(?:\s*[xх×*]\s*\d+)+)\b",
        str(text),
        flags=re.IGNORECASE,
    )
    if multiply_match:
        factors = [int(part) for part in re.split(r"\s*[xх×*]\s*", multiply_match.group(1)) if part.isdigit()]
        if factors:
            product = 1
            for factor in factors:
                product *= factor
            return product, False
        
    # Специальный кейс для Видора микро: упаковки вида "21 шт.+7" и "24 шт.+4"
    # должны учитываться суммарно (21+7=28, 24+4=28), иначе такие позиции
    # ошибочно отбрасываются при сравнении с карточкой товара на сайте (28 шт.).
    if "видора микро" in normalized_text:
        special_match = re.search(
            r"\b(\d+)\s*шт\.?\s*\+\s*(\d+)\b",
            normalized_text,
            flags=re.IGNORECASE,
        )
        if special_match:
            return int(special_match.group(1)) + int(special_match.group(2)), True


    m = re.search(r"(?:\bN\s*|№\s*)([\d+]+)\b", text, flags=re.IGNORECASE)
    if not m:
        m = re.search(
            r"\b(\d+)\s*(?:шт\.?|амп\.?|ампул(?:а|ы)?|шпр\.?|шприц(?:-?тюб)?)\b",
            text,
            flags=re.IGNORECASE,
        )
    if not m:
        return None, False

    raw = m.group(1)
    if "+" in raw:
        parts = [p for p in raw.split("+") if p.isdigit()]
        if not parts:
            return None, True
        return sum(int(p) for p in parts), True
    return int(raw), False


def extract_dosage_from_xls_row(text: str) -> Optional[str]:
    """Возвращает дозировку из текста в нормализованном виде (например, `5 мг + 2 мг`)."""
    if not text:
        return None

    normalized_text = str(text).lower().replace("ё", "е")
    normalized_text = re.sub(r"(?<=\d)\s*[\.,]\s*(?=\d)", ".", normalized_text)
    def _normalize_part(number: float, unit: str) -> tuple[float, str]:
        unit = unit.lower()
        if unit == "мкг":
            return number / 1000, "мг"
        if unit in {"г", "гр"}:
            return number * 1000, "мг"
        return number, unit

    def _format_number(number: float) -> str:
        return (f"{number:.6f}").rstrip("0").rstrip(".")

    def _parentheses_depth(raw_text: str, idx: int) -> int:
        depth = 0
        for pos, ch in enumerate(raw_text):
            if pos >= idx:
                break
            if ch == "(":
                depth += 1
            elif ch == ")" and depth > 0:
                depth -= 1
        return depth

    potency_units = r"ме|мe|me|ед|le|ле|iu"
    matches = list(
        re.finditer(
            rf"\b(\d+(?:[\.,]\d+)?)\s*(мкг|мг|г|мл|{potency_units}|%)(?!\w)",
            normalized_text,
            flags=re.IGNORECASE,
        )
    )
    if not matches:
        return None

    min_depth = min(_parentheses_depth(normalized_text, m.start()) for m in matches)
    selected_matches = [m for m in matches if _parentheses_depth(normalized_text, m.start()) == min_depth]

    parsed_parts: list[tuple[float, str]] = []
    for m in selected_matches:
        raw_number = float(m.group(1))
        number, unit = _normalize_part(raw_number, m.group(2))
        if unit in {"мe", "me", "ед", "le", "ле", "iu"}:
            unit = "ме"
        parsed_parts.append((number, unit))
    parsed_parts = sorted(set(parsed_parts), key=lambda part: (part[1], part[0]))
    parts = [f"{_format_number(number)} {unit}" for number, unit in parsed_parts]

    return " + ".join(parts)